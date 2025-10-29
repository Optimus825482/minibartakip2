from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from flask import session, request
from models import Kullanici, StokHareket, Urun, SistemLog, db
import json


def get_current_user():
    """Oturumdaki kullanıcıyı getir"""
    if 'kullanici_id' in session:
        return Kullanici.query.get(session['kullanici_id'])
    return None


def get_toplam_stok(urun_id):
    """Ürün için toplam stok miktarını hesapla"""
    girisler = db.session.query(db.func.sum(StokHareket.miktar)).filter(
        StokHareket.urun_id == urun_id,
        StokHareket.hareket_tipi.in_(['giris', 'devir', 'sayim'])
    ).scalar() or 0
    
    cikislar = db.session.query(db.func.sum(StokHareket.miktar)).filter(
        StokHareket.urun_id == urun_id,
        StokHareket.hareket_tipi == 'cikis'
    ).scalar() or 0
    
    return girisler - cikislar


def get_kritik_stok_urunler():
    """Kritik stok seviyesinin altındaki ürünleri getir"""
    urunler = Urun.query.filter_by(aktif=True).all()
    kritik_urunler = []
    
    for urun in urunler:
        mevcut_stok = get_toplam_stok(urun.id)
        if mevcut_stok <= urun.kritik_stok_seviyesi:
            kritik_urunler.append({
                'urun': urun,
                'mevcut_stok': mevcut_stok,
                'kritik_seviye': urun.kritik_stok_seviyesi
            })
    
    return kritik_urunler


def excel_export_stok_raporu():
    """Stok durumu Excel raporu oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Raporu"
    
    # Başlık stili
    baslik_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    baslik_fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    baslik_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = ['Ürün Kodu', 'Ürün Adı', 'Grup', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = baslik_alignment
        cell.border = border
    
    # Veri satırları
    urunler = Urun.query.filter_by(aktif=True).all()
    for row_num, urun in enumerate(urunler, 2):
        mevcut_stok = get_toplam_stok(urun.id)
        durum = 'KRİTİK' if mevcut_stok <= urun.kritik_stok_seviyesi else 'NORMAL'
        
        data = [
            urun.id,
            urun.urun_adi,
            urun.grup.grup_adi,
            urun.birim,
            mevcut_stok,
            urun.kritik_stok_seviyesi,
            durum
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if durum == 'KRİTİK' and col_num == 7:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                cell.font = Font(color='DC2626', bold=True)
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def excel_export_zimmet_raporu(personel_id):
    """Personel zimmet Excel raporu oluştur"""
    from models import PersonelZimmet, PersonelZimmetDetay
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Zimmet Raporu"
    
    # Stil tanımlamaları
    baslik_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    baslik_fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    baslik_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = ['Zimmet No', 'Ürün Adı', 'Teslim Edilen', 'Kullanılan', 'Kalan', 'Tarih', 'Durum']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = baslik_alignment
        cell.border = border
    
    # Veri satırları
    zimmetler = PersonelZimmet.query.filter_by(personel_id=personel_id).all()
    row_num = 2
    
    for zimmet in zimmetler:
        for detay in zimmet.detaylar:
            data = [
                zimmet.id,
                detay.urun.urun_adi,
                detay.miktar,
                detay.kullanilan_miktar,
                detay.kalan_miktar or (detay.miktar - detay.kullanilan_miktar),
                zimmet.zimmet_tarihi.strftime('%d.%m.%Y'),
                zimmet.durum.upper()
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            row_num += 1
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def pdf_export_stok_raporu():
    """Stok durumu PDF raporu oluştur"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=30,
        alignment=1
    )
    
    # Başlık
    title = Paragraph("STOK DURUM RAPORU", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Tablo verileri
    data = [['Ürün Kodu', 'Ürün Adı', 'Grup', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']]
    
    urunler = Urun.query.filter_by(aktif=True).all()
    for urun in urunler:
        mevcut_stok = get_toplam_stok(urun.id)
        durum = 'KRİTİK' if mevcut_stok <= urun.kritik_stok_seviyesi else 'NORMAL'
        
        data.append([
            str(urun.id),
            urun.urun_adi[:30],
            urun.grup.grup_adi[:20],
            urun.birim,
            str(mevcut_stok),
            str(urun.kritik_stok_seviyesi),
            durum
        ])
    
    # Tablo oluştur
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#475569')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # PDF oluştur
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


def format_tarih(tarih):
    """Tarih formatla"""
    if not tarih:
        return ''
    return tarih.strftime('%d.%m.%Y %H:%M')


def format_para(tutar):
    """Para formatla"""
    if not tutar:
        return '0,00 ₺'
    return f'{tutar:,.2f} ₺'.replace(',', 'X').replace('.', ',').replace('X', '.')


def log_islem(islem_tipi, modul, islem_detay=None):
    """
    Sistem işlemlerini logla
    
    Args:
        islem_tipi: giris, cikis, ekleme, guncelleme, silme, goruntuleme
        modul: urun, stok, zimmet, oda, kat, personel vb.
        islem_detay: İşlem detayları (dict veya string)
    """
    try:
        kullanici_id = session.get('kullanici_id')
        
        # İşlem detayını JSON'a çevir
        if isinstance(islem_detay, dict):
            detay_json = json.dumps(islem_detay, ensure_ascii=False)
        else:
            detay_json = str(islem_detay) if islem_detay else None
        
        # IP adresi ve tarayıcı bilgisi
        ip_adresi = request.remote_addr if request else None
        tarayici = request.headers.get('User-Agent', '')[:200] if request else None
        
        # Log kaydı oluştur
        log = SistemLog(
            kullanici_id=kullanici_id,
            islem_tipi=islem_tipi,
            modul=modul,
            islem_detay=detay_json,
            ip_adresi=ip_adresi,
            tarayici=tarayici
        )
        
        db.session.add(log)
        db.session.commit()
        
    except Exception as e:
        # Log hatası uygulamayı durdurmamalı
        print(f'Log hatası: {str(e)}')
        try:
            db.session.rollback()
        except:
            pass


def get_son_loglar(limit=50):
    """Son log kayıtlarını getir"""
    return SistemLog.query.order_by(SistemLog.islem_tarihi.desc()).limit(limit).all()


def get_kullanici_loglari(kullanici_id, limit=50):
    """Belirli bir kullanıcının log kayıtlarını getir"""
    return SistemLog.query.filter_by(kullanici_id=kullanici_id).order_by(
        SistemLog.islem_tarihi.desc()
    ).limit(limit).all()


def get_modul_loglari(modul, limit=50):
    """Belirli bir modülün log kayıtlarını getir"""
    return SistemLog.query.filter_by(modul=modul).order_by(
        SistemLog.islem_tarihi.desc()
    ).limit(limit).all()

